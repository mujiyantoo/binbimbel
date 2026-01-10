from fastapi import APIRouter, HTTPException, status, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc
from models.registration import (
    RegistrationCreate, 
    RegistrationModel, 
    RegistrationResponse,
    RegistrationListResponse
)
from database import get_db
from typing import List
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/registrations", tags=["registrations"])


@router.post("", response_model=RegistrationResponse, status_code=status.HTTP_201_CREATED)
async def create_registration(registration: RegistrationCreate, db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk submit formulir pendaftaran siswa baru
    """
    try:
        # Konversi ke model SQLAlchemy
        new_reg = RegistrationModel(**registration.dict())
        
        # Simpan ke Database
        db.add(new_reg)
        await db.commit()
        await db.refresh(new_reg)
        
        logger.info(f"New registration created: {new_reg.id}")
        
        return RegistrationResponse(
            success=True,
            message="Pendaftaran berhasil! Tim kami akan segera menghubungi Anda.",
            data={
                "registration_id": new_reg.id,
                "nama_lengkap": new_reg.nama_lengkap,
                "program": new_reg.program,
                "created_at": new_reg.created_at.isoformat()
            }
        )
        
    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error creating registration: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat memproses pendaftaran. Silakan coba lagi."
        )


@router.get("", response_model=RegistrationListResponse)
async def get_all_registrations(
    skip: int = 0,
    limit: int = 100,
    program: str = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk mengambil semua data pendaftaran
    """
    try:
        # Build query
        stmt = select(RegistrationModel)
        if program:
            stmt = stmt.where(RegistrationModel.program == program)
        
        # Sorting, pagination
        stmt = stmt.order_by(desc(RegistrationModel.created_at)).offset(skip).limit(limit)
        
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        
        formatted_data = []
        for reg in registrations:
            formatted_data.append({
                "registration_id": reg.id,
                "nama_lengkap": reg.nama_lengkap,
                "program": reg.program,
                "telepon": reg.telepon,
                "kelas": reg.kelas,
                "created_at": reg.created_at.isoformat() if reg.created_at else ""
            })
        
        return RegistrationListResponse(
            success=True,
            count=len(formatted_data),
            data=formatted_data
        )
        
    except Exception as e:
        logger.error(f"Error fetching registrations: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat mengambil data"
        )


@router.get("/{registration_id}", response_model=RegistrationResponse)
async def get_registration_by_id(registration_id: str, db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mengambil detail satu pendaftaran berdasarkan ID
    """
    try:
        stmt = select(RegistrationModel).where(RegistrationModel.id == registration_id)
        result = await db.execute(stmt)
        registration = result.scalar_one_or_none()
        
        if not registration:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Registration not found"
            )
        
        # Convert to dict for response
        reg_dict = {
            "id": registration.id,
            "nama_lengkap": registration.nama_lengkap,
            "nama_panggilan": registration.nama_panggilan,
            "jenis_kelamin": registration.jenis_kelamin,
            "tempat_lahir": registration.tempat_lahir,
            "tanggal_lahir": registration.tanggal_lahir,
            "asal_sekolah": registration.asal_sekolah,
            "kelas": registration.kelas,
            "alamat": registration.alamat,
            "telepon": registration.telepon,
            "email": registration.email,
            "nama_ayah": registration.nama_ayah,
            "pekerjaan_ayah": registration.pekerjaan_ayah,
            "telepon_ayah": registration.telepon_ayah,
            "nama_ibu": registration.nama_ibu,
            "pekerjaan_ibu": registration.pekerjaan_ibu,
            "telepon_ibu": registration.telepon_ibu,
            "alamat_ortu": registration.alamat_ortu,
            "program": registration.program,
            "mata_pelajaran": registration.mata_pelajaran,
            "hari": registration.hari,
            "waktu": registration.waktu,
            "referensi": registration.referensi,
            "persetujuan": registration.persetujuan,
            "tanggal_daftar": registration.tanggal_daftar,
            "created_at": registration.created_at.isoformat() if registration.created_at else None,
            "updated_at": registration.updated_at.isoformat() if registration.updated_at else None
        }
        
        return RegistrationResponse(
            success=True,
            message="Data ditemukan",
            data=reg_dict
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching registration {registration_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat mengambil data"
        )


@router.get("/stats/summary")
async def get_registration_stats(db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mendapatkan statistik pendaftaran
    """
    try:
        # Total registrations
        total_stmt = select(func.count(RegistrationModel.id))
        total_result = await db.execute(total_stmt)
        total = total_result.scalar()
        
        # Count by program
        program_stmt = select(RegistrationModel.program, func.count(RegistrationModel.id)).group_by(RegistrationModel.program)
        program_result = await db.execute(program_stmt)
        program_stats = {row[0]: row[1] for row in program_result.all()}
        
        # Count by level (Substr logic)
        # Note: SUBSTR works in most Postgres versions
        level_stmt = select(func.substr(RegistrationModel.kelas, 1, 2), func.count(RegistrationModel.id)).group_by(func.substr(RegistrationModel.kelas, 1, 2))
        level_result = await db.execute(level_stmt)
        level_stats = {row[0]: row[1] for row in level_result.all()}
        
        return {
            "success": True,
            "data": {
                "total_registrations": total,
                "by_program": program_stats,
                "by_level": level_stats
            }
        }
        
    except Exception as e:
        logger.error(f"Error fetching stats: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat mengambil statistik"
        )


@router.get("/export/excel")
async def export_registrations_excel(db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk export data pendaftaran ke Excel
    """
    try:
        # Fetch all registrations
        stmt = select(RegistrationModel).order_by(desc(RegistrationModel.created_at))
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        
        if not registrations:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Tidak ada data untuk diexport"
            )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pendaftaran"
        
        # Header style
        header_fill = PatternFill(start_color="5A9C9B", end_color="5A9C9B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Header row
        headers = [
            "No",
            "Tanggal Daftar",
            "Nama Lengkap",
            "Nama Panggilan",
            "Jenis Kelamin",
            "Tempat Lahir",
            "Tanggal Lahir",
            "Asal Sekolah",
            "Kelas",
            "Alamat",
            "Telepon",
            "Email",
            "Nama Ayah",
            "Telepon Ayah",
            "Alamat Orang Tua",
            "Program",
            "Mata Pelajaran",
            "Hari",
            "Waktu",
            "Referensi"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, reg in enumerate(registrations, 2):
            # Format mata pelajaran
            mata_pelajaran_str = ", ".join(reg.mata_pelajaran) if reg.mata_pelajaran else ""
            
            # Format tanggal
            created_at_str = reg.created_at.strftime("%Y-%m-%d %H:%M:%S") if reg.created_at else ""
            tanggal_daftar_str = reg.tanggal_daftar if reg.tanggal_daftar else ""
            
            # Jenis kelamin
            jenis_kelamin_str = "Laki-laki" if reg.jenis_kelamin == "L" else "Perempuan" if reg.jenis_kelamin == "P" else ""
            
            row_data = [
                row_num - 1,  # No
                tanggal_daftar_str,
                reg.nama_lengkap or "",
                reg.nama_panggilan or "",
                jenis_kelamin_str,
                reg.tempat_lahir or "",
                reg.tanggal_lahir or "",
                reg.asal_sekolah or "",
                reg.kelas.upper() if reg.kelas else "",
                reg.alamat or "",
                reg.telepon or "",
                reg.email or "",
                reg.nama_ayah or "",
                reg.telepon_ayah or "",
                reg.alamat_ortu or "",
                reg.program or "",
                mata_pelajaran_str,
                reg.hari or "",
                reg.waktu or "",
                reg.referensi or ""
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with current date
        filename = f"Data_Pendaftaran_BIN_Bimbel_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat export Excel"
        )
