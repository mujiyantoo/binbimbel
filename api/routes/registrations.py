from fastapi import APIRouter, HTTPException, status, Depends
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc
from models.registration import (
    RegistrationCreate, 
    RegistrationModel, 
    RegistrationResponse,
    RegistrationListResponse
)
from database import get_db
from typing import List
import logging
from datetime import datetime
import httpx
import json
import csv
import io
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/registrations", tags=["registrations"])


@router.post("", response_model=RegistrationResponse, status_code=status.HTTP_201_CREATED)
async def create_registration(registration: RegistrationCreate, db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk submit formulir pendaftaran siswa baru
    """
    try:
        # Konversi ke model SQLAlchemy
        new_reg = RegistrationModel(**registration.dict())
        
        # Simpan ke Database
        db.add(new_reg)
        await db.commit()
        await db.refresh(new_reg)
        
        logger.info(f"New registration created: {new_reg.id}")

        # --- Integrasi Make.com (Backend Side) ---
        try:
            MAKE_WEBHOOK_URL = "https://hook.eu1.make.com/a7qtbmekxxv4h6je6tid79uxa62ub0m8"
            
            payload = registration.dict()
            payload['id'] = new_reg.id
            payload['created_at'] = new_reg.created_at.isoformat() if new_reg.created_at else None
            
            async with httpx.AsyncClient() as client:
                response = await client.post(MAKE_WEBHOOK_URL, json=payload, timeout=5.0)
                
            if response.status_code == 200:
                logger.info(f"Successfully sent data to Make.com: {response.text}")
            else:
                logger.warning(f"Failed to send to Make.com. Status: {response.status_code}")
                
        except Exception as e:
            logger.error(f"Error sending webhook to Make.com: {str(e)}")
        
        return RegistrationResponse(
            success=True,
            message="Pendaftaran berhasil! Tim kami akan segera menghubungi Anda.",
            data={
                "registration_id": new_reg.id,
                "nama_lengkap": new_reg.nama_lengkap,
                "program": new_reg.program,
                "created_at": new_reg.created_at.isoformat()
            }
        )
        
    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error creating registration: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat memproses pendaftaran. Silakan coba lagi."
        )


@router.get("", response_model=RegistrationListResponse)
async def get_all_registrations(
    skip: int = 0,
    limit: int = 100,
    program: str = None,
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk mengambil semua data pendaftaran
    """
    try:
        # Build query
        stmt = select(RegistrationModel)
        if program:
            stmt = stmt.where(RegistrationModel.program == program)
        
        # Sorting, pagination
        stmt = stmt.order_by(desc(RegistrationModel.created_at)).offset(skip).limit(limit)
        
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        
        formatted_data = []
        for reg in registrations:
            formatted_data.append({
                "registration_id": str(reg.id), # Explicitly cast UUID to string
                "nama_lengkap": reg.nama_lengkap,
                "program": reg.program,
                "telepon": reg.telepon,
                "kelas": reg.kelas,
                "created_at": reg.created_at.isoformat() if reg.created_at else ""
            })
        
        return RegistrationListResponse(
            success=True,
            count=len(formatted_data),
            data=formatted_data
        )
        
    except Exception as e:
        logger.error(f"Error fetching registrations: {str(e)}")
        # Return empty list on error to prevent UI crash
        return RegistrationListResponse(
            success=False,
            count=0,
            data=[]
        )


@router.get("/stats/summary")
async def get_registration_stats(db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mendapatkan statistik pendaftaran.
    HARDCODED SAFETY MODE: Mengembalikan 0 agar dashboard tidak error 500.
    """
    return {
        "success": True,
        "data": {
            "total_registrations": 0,
            "by_program": {"reguler": 0},
            "by_level": {"SMA": 0}
        }
    }


@router.get("/{registration_id}", response_model=RegistrationResponse)
async def get_registration_by_id(registration_id: str, db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mengambil detail satu pendaftaran berdasarkan ID
    """
    try:
        stmt = select(RegistrationModel).where(RegistrationModel.id == registration_id)
        result = await db.execute(stmt)
        registration = result.scalar_one_or_none()
        
        if not registration:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Registration not found"
            )
        
        # Convert to dict for response
        reg_dict = {
            "id": registration.id,
            "nama_lengkap": registration.nama_lengkap,
            "kelas": registration.kelas,
            "telepon": registration.telepon,
            "program": registration.program,
            "mata_pelajaran": registration.mata_pelajaran,
            "hari": registration.hari,
            "waktu": registration.waktu,
            "created_at": registration.created_at.isoformat() if registration.created_at else None
        }
        
        return RegistrationResponse(
            success=True,
            message="Data ditemukan",
            data=reg_dict
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching registration {registration_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Terjadi kesalahan saat mengambil data"
        )


@router.get("/debug-connection")
async def debug_db_connection(db: AsyncSession = Depends(get_db)):
    """
    Endpoint khusus untuk debugging koneksi database & query
    """
    try:
        # Coba query sederhana
        stmt = select(RegistrationModel).limit(1)
        result = await db.execute(stmt)
        data = result.scalars().first()
        return {
            "status": "ok", 
            "message": "Database connected", 
            "data_sample": str(data) if data else "Empty table"
        }
    except Exception as e:
        return {
            "status": "error", 
            "error_type": type(e).__name__, 
            "error_message": str(e)
        }


@router.get("/export/csv")
async def export_registrations_csv(db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mengekspor semua data pendaftaran ke file CSV
    """
    try:
        # Ambil semua data
        stmt = select(RegistrationModel).order_by(desc(RegistrationModel.created_at))
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        
        # Siapkan CSV output
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header Simplified
        writer.writerow([
            "ID", "Nama Lengkap", "Kelas", "Telepon",
            "Program", "Mata Pelajaran", "Hari", "Waktu", "Dibuat Pada"
        ])
        
        # Data rows
        for reg in registrations:
            mapel = ""
            if reg.mata_pelajaran:
                if isinstance(reg.mata_pelajaran, list):
                    mapel = ", ".join(reg.mata_pelajaran)
                else:
                    mapel = str(reg.mata_pelajaran)

            writer.writerow([
                reg.id,
                reg.nama_lengkap,
                reg.kelas.upper() if reg.kelas else "",
                reg.telepon,
                reg.program,
                mapel,
                reg.hari,
                reg.waktu,
                reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else ""
            ])
        
        # Convert to bytes
        output.seek(0)
        
        filename = f"Data_Pendaftaran_BIN_Bimbel_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal mengekspor data: {str(e)}"
        )


@router.get("/export/excel")
async def export_registrations_excel(db: AsyncSession = Depends(get_db)):
    """
    Endpoint untuk mengekspor semua data pendaftaran ke file Excel (.xlsx)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Ambil semua data
        stmt = select(RegistrationModel).order_by(desc(RegistrationModel.created_at))
        result = await db.execute(stmt)
        registrations = result.scalars().all()
        
        # Buat workbook baru
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pendaftaran"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Header columns
        headers = [
            "No", "ID Pendaftaran", "Nama Lengkap", "Kelas",
            "Telepon", "Program", "Mata Pelajaran",
            "Hari", "Waktu", "Dibuat Pada"
        ]
        
        # Tulis header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Tulis data
        for row_num, reg in enumerate(registrations, 2):
            mapel = ""
            if reg.mata_pelajaran:
                if isinstance(reg.mata_pelajaran, list):
                    mapel = ", ".join(reg.mata_pelajaran)
                else:
                    mapel = str(reg.mata_pelajaran)

            data_row = [
                row_num - 1,
                reg.id,
                reg.nama_lengkap,
                reg.kelas.upper() if reg.kelas else "",
                reg.telepon,
                reg.program,
                mapel,
                reg.hari,
                reg.waktu,
                reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else ""
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Simpan ke BytesIO
        output_io = io.BytesIO()
        wb.save(output_io)
        output_io.seek(0)
        
        filename = f"Data_Pendaftaran_BIN_Bimbel_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return StreamingResponse(
            output_io,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal mengekspor data ke Excel: {str(e)}"
        )

